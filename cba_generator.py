# cba_generator.py
from __future__ import annotations

import io
import re
from datetime import date
from io import BytesIO
from typing import Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def safe_name(s: str) -> str:
    return re.sub(r'[\\/*?:<>|"]+', "", s)


def generate_cba_from_uploaded_template(
    uploaded_xlsx_bytes: bytes,
    purpose: str,
    project_name: str,
    project_location: str,
    sheet_name: Optional[str] = None,
) -> Tuple[bytes, str]:
    # ---------- CONFIG ----------
    TITLE = f"Choose-by-Advantage Matrix for the {purpose}"
    SECTION_BAND_FILL = "f5f5f5"
    ROWLABEL_FILL = "f0f0f0"
    BORDER_COLOR = "000000"
    TITLE_SIZE = 18
    HEADER_FONT_SIZE = 12
    SECTION_TEXT_ROTATE = 90

    RATING_WORDS = ["Poor", "Fair", "Good", "Very Good", "Excellent"]
    RATING_COLOR = {
        "Poor": "FFC000",
        "Fair": "FFFF66",
        "Good": "CCFF66",
        "Very Good": "78FE66",
        "Excellent": "19CB01",
    }

    WORD_TO_NUM = {"poor": 1, "fair": 2, "good": 3, "very good": 4, "excellent": 5}

    DESC_ROWS = {"illustration", "description", "feasibility", "advantages", "disadvantages", "scheme"}

    CONSID_ROWS = {
        "foundation installation schedule",
        "installation schedule",
        "equipment/subcontractors necessary for foundations",
        "equipment/subcontractors",
        "spoils handling",
        "certainty of improvement",
        "authority having jurisdiction approval",
        "noise",
        "vibration",
        "cost",
        "market competition",
        "market familiarity",
    }

    NO_BULLET_ROWS = {"feasibility"}  # keep feasibility plain

    # ---------- helpers ----------
    def bulletize(text):
        if text is None:
            return ""
        s = str(text).strip()
        if not s:
            return ""
        if any(sep in s for sep in ["\n", ";", " • ", " - ", " – "]):
            s = s.replace("\r", "")
            for sep in [";", " • ", " – ", " - "]:
                s = s.replace(sep, "\n")
            lines = [re.sub(r"^\s*[-•–]\s*", "", ln).strip() for ln in s.split("\n")]
            lines = [ln for ln in lines if ln]
            if lines:
                return "• " + "\n• ".join(lines)
        return s

    _RATING_RE = re.compile(
        r"^\s*(very\s+good|excellent|good|fair|poor|\d(?:\.0)?)\s*(?:[:\-–—]\s*(.*))?$",
        re.I,
    )

    def split_rating_and_desc(raw):
        if raw is None or str(raw).strip() == "":
            return "", ""
        s = str(raw).strip()
        m = _RATING_RE.match(s)
        if not m:
            return "", s
        rt = m.group(1).lower()
        if rt in WORD_TO_NUM:
            rating_word = "Very Good" if rt == "very good" else rt.title()
        else:
            try:
                rating_word = ["", "Poor", "Fair", "Good", "Very Good", "Excellent"][int(float(rt))]
            except Exception:
                rating_word = ""
        return rating_word, (m.group(2) or "")

    # ---------- read Excel template ----------
    xls = pd.ExcelFile(io.BytesIO(uploaded_xlsx_bytes), engine="openpyxl")
    sheet_to_use = xls.sheet_names[0] if sheet_name is None else sheet_name
    df0 = pd.read_excel(xls, sheet_name=sheet_to_use, engine="openpyxl")

    df0.columns = [str(c).strip() for c in df0.columns]
    row_label_header = df0.columns[0]
    options = list(df0.columns[1:])
    labels = [str(x) for x in df0.iloc[:, 0].tolist()]
    lower = [x.strip().lower() for x in labels]

    # Ensure there is an Illustration row
    if "illustration" not in lower:
        df0 = pd.concat(
            [pd.DataFrame([["Illustration"] + [""] * len(options)], columns=df0.columns), df0],
            ignore_index=True,
        )
        labels = [str(x) for x in df0.iloc[:, 0].tolist()]
        lower = [x.strip().lower() for x in labels]

    # =========================
    #   Matrix sheet
    # =========================
    wb = Workbook()
    ws = wb.active
    ws.title = "Matrix"

    THICK = Side(style="thick", color=BORDER_COLOR)
    THIN = Side(style="thin", color=BORDER_COLOR)
    MEDIUM = Side(style="medium", color=BORDER_COLOR)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
    VERT_CENTER = Alignment(
        horizontal="center", vertical="center", textRotation=SECTION_TEXT_ROTATE, wrap_text=True
    )

    ncols = 2 + len(options)

    # Banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=TITLE)
    t.font = Font(bold=True, size=TITLE_SIZE)
    t.alignment = Alignment(horizontal="center", vertical="center")

    info_r = 2
    left_end = max(2, int(ncols * 0.25))
    right_start = max(left_end + 1, ncols - max(2, int(ncols * 0.25)) + 1)
    mid_start = left_end + 1
    mid_end = right_start - 1

    ws.merge_cells(start_row=info_r, start_column=1, end_row=info_r, end_column=left_end)
    ws.merge_cells(start_row=info_r, start_column=mid_start, end_row=info_r, end_column=mid_end)
    ws.merge_cells(start_row=info_r, start_column=right_start, end_row=info_r, end_column=ncols)

    ws.cell(row=info_r, column=1, value=f"Project Name: {project_name}").font = Font(bold=True, size=14)
    cC = ws.cell(row=info_r, column=mid_start, value=f"Project Location: {project_location}")
    cC.font = Font(bold=True, size=14)
    cC.alignment = Alignment(horizontal="center", vertical="center")
    cR = ws.cell(row=info_r, column=right_start, value=f"Date: {date.today():%B %d, %Y}")
    cR.font = Font(bold=True, size=14)
    cR.alignment = Alignment(horizontal="right", vertical="center")

    # Header row
    hr = 4

    def style_hdr(r, c, text=None):
        cell = ws.cell(row=r, column=c, value=text if text else None)
        if text:
            cell.font = Font(bold=True, size=HEADER_FONT_SIZE)
        cell.alignment = CENTER
        return cell

    style_hdr(hr, 1, text=None)
    display_header = "Options" if "unnamed" in row_label_header.lower() else row_label_header
    style_hdr(hr, 2, display_header)
    for idx, name in enumerate(options, start=1):
        style_hdr(hr, 2 + idx, f"Option {idx} - {name}")

    # Section grouping
    desc_idx, consid_idx, other_idx = [], [], []
    for i, lab in enumerate(lower):
        if lab in DESC_ROWS:
            desc_idx.append(i)
        elif lab in CONSID_ROWS:
            consid_idx.append(i)
        else:
            other_idx.append(i)

    first_section_name = f"{purpose} Description"
    order = [(first_section_name, desc_idx), ("Construction Considerations", consid_idx + other_idx)]

    r = hr + 1
    upper_cells_by_col = {j: [] for j in range(1, len(options) + 1)}
    attr_names_cc = []

    for section_name, idxs in order:
        if not idxs:
            continue
        start_r_section = r

        for i in idxs:
            label = labels[i]
            is_cc = section_name == "Construction Considerations"

            if is_cc:
                upper_r, lower_r = r, r + 1
                attr_names_cc.append(label)

                ws.merge_cells(start_row=upper_r, start_column=2, end_row=lower_r, end_column=2)
                labcell = ws.cell(row=upper_r, column=2, value=label)
                labcell.font = Font(bold=True)
                labcell.fill = PatternFill("solid", fgColor=ROWLABEL_FILL)
                labcell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                for j in range(1, len(options) + 1):
                    df_col = 1 + (j - 1)
                    excel_col = 2 + j
                    raw = df0.iat[i, df_col] if df_col < df0.shape[1] else ""
                    rating_word, desc = split_rating_and_desc(raw)

                    ws.cell(row=upper_r, column=excel_col, value=rating_word).alignment = CENTER
                    upper_cells_by_col[j].append(f"{get_column_letter(excel_col)}{upper_r}")

                    ws.cell(row=lower_r, column=excel_col, value=bulletize(desc)).alignment = LEFT

                ws.row_dimensions[upper_r].height = 15
                r += 2

            else:
                ws.cell(row=r, column=2, value=label).alignment = LEFT
                ws.cell(row=r, column=2).font = Font(bold=True)
                ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=ROWLABEL_FILL)

                for j in range(1, len(options) + 1):
                    df_col = 1 + (j - 1)
                    excel_col = 2 + j
                    raw = df0.iat[i, df_col] if df_col < df0.shape[1] else ""

                    rw, desc = split_rating_and_desc(raw)
                    base_text = str(raw or "")

                    if label.strip().lower() in NO_BULLET_ROWS:
                        txt, align = base_text, LEFT
                    else:
                        txt = rw if (desc == "" and rw) else bulletize(desc or base_text)
                        align = CENTER if (desc == "" and rw) else LEFT

                    ws.cell(row=r, column=excel_col, value=txt).alignment = align
                r += 1

        end_r_section = r - 1
        ws.merge_cells(start_row=start_r_section, start_column=1, end_row=end_r_section, end_column=1)
        band = ws.cell(row=start_r_section, column=1, value=section_name)
        band.fill = PatternFill("solid", fgColor=SECTION_BAND_FILL)
        band.font = Font(bold=True)
        band.alignment = VERT_CENTER

    table_end_row = r - 1

    # Borders: thick frame, medium vertical between options, thick horizontal under Disadvantages
    disadv_row_top = None
    try:
        disadv_idx = next((i for i, lab in enumerate(lower) if lab.strip().lower() == "disadvantages"))
    except StopIteration:
        disadv_idx = None

    if disadv_idx is not None:
        for row_num in range(hr + 1, table_end_row + 1):
            if str(ws.cell(row=row_num, column=2).value).strip() == str(labels[disadv_idx]).strip():
                disadv_row_top = row_num
                break

    for rr in range(hr, table_end_row + 1):
        for cc in range(1, 2 + len(options) + 1):
            left_b = THICK if cc == 1 else THIN
            right_b = THICK if cc == (2 + len(options)) else THIN
            top_b = THICK if rr == hr else THIN
            bottom_b = THICK if rr == table_end_row else THIN

            if cc >= 3 and cc < (2 + len(options)):
                left_b = MEDIUM  # option dividers

            if disadv_row_top and rr == disadv_row_top + 1:
                top_b = THICK

            ws.cell(row=rr, column=cc).border = Border(left=left_b, right=right_b, top=top_b, bottom=bottom_b)

    ws.column_dimensions["A"].width = 5
    for col in range(2, 2 + len(options) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 35

    # DV + CF (upper rating cells)
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(RATING_WORDS) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid rating",
        error="Choose a rating from the list.",
        promptTitle="Select rating",
        prompt="Pick a rating.",
    )
    ws.add_data_validation(dv)

    for j in range(1, len(options) + 1):
        for addr in upper_cells_by_col[j]:
            dv.add(addr)
            for word, hexcol in RATING_COLOR.items():
                dxf = DifferentialStyle(fill=PatternFill(fill_type="solid", start_color=hexcol, end_color=hexcol))
                rule = Rule(type="cellIs", operator="equal", formula=[f'"{word}"'], dxf=dxf)
                ws.conditional_formatting.add(addr, rule)

    # Freeze & print
    ws.freeze_panes = "C5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "4:4"

    # =========================
    #   Weights & SAW
    # =========================
    wsw = wb.create_sheet("Weights & SAW")

    opt_headers = [f"NormScore - {opt}" for opt in options]
    headers = ["Attribute", "Active?", "Importance (1–5)", "Weight (normalized)"] + opt_headers
    wsw.append(headers)

    for c in range(1, len(headers) + 1):
        h = wsw.cell(row=1, column=c)
        h.font = Font(bold=True)
        h.alignment = CENTER

    # Rating map at L/M
    map_col = 12  # L/M
    wsw.cell(row=1, column=map_col, value="Rating").font = Font(bold=True)
    wsw.cell(row=1, column=map_col + 1, value="Raw").font = Font(bold=True)
    for r0, (name, val) in enumerate(
        [("Poor", 1), ("Fair", 2), ("Good", 3), ("Very Good", 4), ("Excellent", 5)], start=2
    ):
        wsw.cell(row=r0, column=map_col, value=name)
        wsw.cell(row=r0, column=map_col + 1, value=val)

    dv_active = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv_import = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=False)
    wsw.add_data_validation(dv_active)
    wsw.add_data_validation(dv_import)

    n_attr = len(attr_names_cc)
    default_importance = 3
    active_rng = f"$B$2:$B${1 + n_attr}"
    imp_rng = f"$C$2:$C${1 + n_attr}"
    w_rng = f"$D$2:$D${1 + n_attr}"

    for i, attr in enumerate(attr_names_cc, start=2):
        wsw.cell(row=i, column=1, value=attr)

        wsw.cell(row=i, column=2, value="Yes")
        dv_active.add(f"B{i}")

        wsw.cell(row=i, column=3, value=default_importance)
        dv_import.add(f"C{i}")

        # Weight normalization a_i / sum(a_i) (active only)
        wsw.cell(
            row=i,
            column=4,
            value=(
                f'=IF(SUMPRODUCT(({active_rng}="Yes")*({imp_rng}))=0,0,'
                f'IF(B{i}="Yes",C{i},0)/SUMPRODUCT(({active_rng}="Yes")*({imp_rng})))'
            ),
        )

        # NormScores: INDEX/MATCH, clamped
        for j in range(1, len(options) + 1):
            col_idx = 4 + j  # E...
            addr_rating = upper_cells_by_col[j][i - 2]  # rating cell address in Matrix sheet
            mapR = get_column_letter(map_col)
            mapV = get_column_letter(map_col + 1)
            formula = (
                f"=MAX(0,(IFERROR(INDEX(${mapV}$2:${mapV}$6,"
                f"MATCH('Matrix'!{addr_rating},${mapR}$2:${mapR}$6,0)),0)-1)/4)"
            )
            wsw.cell(row=i, column=col_idx, value=formula)

    # SAW / Rank / Score(0–10) / SumW
    total_row = 2 + n_attr
    wsw.cell(row=total_row, column=1, value="SAW Score")

    first_norm_col = 5
    last_norm_col = 4 + len(options)

    for j in range(1, len(options) + 1):
        col_idx = 4 + j
        total_formula = f"=SUMPRODUCT({w_rng},{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{1 + n_attr})"
        wsw.cell(row=total_row, column=col_idx, value=total_formula)

    rank_row2 = total_row + 1
    wsw.cell(row=rank_row2, column=1, value="Rank")
    rank_range = f"${get_column_letter(first_norm_col)}${total_row}:${get_column_letter(last_norm_col)}${total_row}"
    for j in range(first_norm_col, last_norm_col + 1):
        saw_cell = f"{get_column_letter(j)}{total_row}"
        wsw.cell(row=rank_row2, column=j, value=f"=RANK({saw_cell},{rank_range},0)")

    score10_row = total_row + 2
    wsw.cell(row=score10_row, column=1, value="Score (0–10)")
    for j in range(first_norm_col, last_norm_col + 1):
        saw_cell = f"{get_column_letter(j)}{total_row}"
        wsw.cell(row=score10_row, column=j, value=f"=ROUND(10*{saw_cell},0)")

    sumw_row = score10_row + 1
    wsw.cell(row=sumw_row, column=1, value="Sum of normalized weights")
    wsw.cell(row=sumw_row, column=2, value=f"=SUM({w_rng})")

    # formats + alignment
    for rr in range(2, 2 + n_attr):
        wsw.cell(row=rr, column=4).number_format = "0.00"
    for j in range(first_norm_col, last_norm_col + 1):
        wsw.cell(row=total_row, column=j).number_format = "0.00"

    max_col = last_norm_col
    max_row = sumw_row
    for rr in range(1, max_row + 1):
        for cc in range(1, max_col + 1):
            wsw.cell(row=rr, column=cc).alignment = CENTER

    widths = [34, 10, 18, 18] + [18] * len(options)
    for c, wid in enumerate(widths, start=1):
        wsw.column_dimensions[get_column_letter(c)].width = wid
    wsw.column_dimensions[get_column_letter(map_col)].width = 12
    wsw.column_dimensions[get_column_letter(map_col + 1)].width = 9
    wsw.freeze_panes = "A2"

    # Header live color based on SAW*10 (Matrix headers)
    bands = [
        (0, 2, "FFC000"),
        (2, 4, "FFFF66"),
        (4, 6, "CCFF66"),
        (6, 8, "78FE66"),
        (8, 11, "19CB01"),
    ]
    for j in range(1, len(options) + 1):
        hdr_col = 2 + j
        hdr_cell = f"{get_column_letter(hdr_col)}{hr}"
        saw_cell = f"'Weights & SAW'!{get_column_letter(4 + j)}${total_row}"  # SAW (0..1)
        for lo, hi, hexcol in bands:
            expr = f"=AND(10*{saw_cell}>={lo},10*{saw_cell}<{hi})"
            dxf = DifferentialStyle(fill=PatternFill(fill_type="solid", start_color=hexcol, end_color=hexcol))
            rule = Rule(type="expression", dxf=dxf, formula=[expr])
            ws.conditional_formatting.add(hdr_cell, rule)

    # =========================
    #   Summary CBA sheet
    # =========================
    ws_sum = wb.create_sheet("Summary CBA")

    CENTER_SUM = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_SUM = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ncols_sum = 1 + len(options)  # A = labels, B.. = options

    # Title row (row 1)
    ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_sum)
    t2 = ws_sum.cell(row=1, column=1, value=TITLE)
    t2.font = Font(bold=True, size=TITLE_SIZE)
    t2.alignment = Alignment(horizontal="center", vertical="center")

    # Compute Summary header merge geometry from Summary width
    left_end_s = max(2, int(ncols_sum * 0.25))
    right_start_s = max(left_end_s + 1, ncols_sum - max(2, int(ncols_sum * 0.25)) + 1)
    mid_start_s = left_end_s + 1
    mid_end_s = right_start_s - 1

    ws_sum.merge_cells(start_row=2, start_column=1, end_row=2, end_column=left_end_s)
    ws_sum.merge_cells(start_row=2, start_column=mid_start_s, end_row=2, end_column=mid_end_s)
    ws_sum.merge_cells(start_row=2, start_column=right_start_s, end_row=2, end_column=ncols_sum)

    ws_sum.cell(row=2, column=1, value=f"Project Name: {project_name}").font = Font(bold=True, size=14)
    c_mid = ws_sum.cell(row=2, column=mid_start_s, value=f"Project Location: {project_location}")
    c_mid.font = Font(bold=True, size=14)
    c_mid.alignment = Alignment(horizontal="center", vertical="center")
    c_rt = ws_sum.cell(row=2, column=right_start_s, value=f"Date: {date.today():%B %d, %Y}")
    c_rt.font = Font(bold=True, size=14)
    c_rt.alignment = Alignment(horizontal="right", vertical="center")

    # Leave row 3 as spacer
    row_ill, row_opt, row_desc, row_score, row_summary = 4, 5, 6, 7, 8

    # Labels in column A
    labels_summary = ["Illustration", "Option", "Description", "Score", "Summary"]
    for rr, lab in zip([row_ill, row_opt, row_desc, row_score, row_summary], labels_summary):
        cell = ws_sum.cell(row=rr, column=1, value=lab)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=ROWLABEL_FILL)

    def find_matrix_row(label_lower: str):
        for rr in range(hr + 1, table_end_row + 1):
            val = ws.cell(row=rr, column=2).value
            if str(val).strip().lower() == label_lower:
                return rr
        return None

    # Illustration (text reference only; images cannot be referenced by formula)
    matrix_row_ill = find_matrix_row("illustration")
    if matrix_row_ill:
        for j in range(1, len(options) + 1):
            ws_sum.cell(
                row=row_ill,
                column=1 + j,
                value=f"='Matrix'!{get_column_letter(2 + j)}{matrix_row_ill}",
            ).alignment = CENTER_SUM

    # Option headers
    for j in range(1, len(options) + 1):
        c = ws_sum.cell(row=row_opt, column=1 + j, value=f"='Matrix'!{get_column_letter(2 + j)}{hr}")
        c.font = Font(bold=True)
        c.alignment = CENTER_SUM

    # Apply same SAW-score color bands to Summary option row
    for j in range(1, len(options) + 1):
        hdr_cell = f"{get_column_letter(1 + j)}{row_opt}"
        saw_cell = f"'Weights & SAW'!{get_column_letter(4 + j)}{total_row}"
        for lo, hi, hexcol in bands:
            expr = f"=AND(10*{saw_cell}>={lo},10*{saw_cell}<{hi})"
            dxf = DifferentialStyle(fill=PatternFill(fill_type="solid", start_color=hexcol, end_color=hexcol))
            rule = Rule(type="expression", dxf=dxf, formula=[expr])
            ws_sum.conditional_formatting.add(hdr_cell, rule)

    # Description row: prefer Scheme; else Description
    use_label = "scheme" if "scheme" in lower else ("description" if "description" in lower else None)
    matrix_row_desc = find_matrix_row(use_label) if use_label else None
    if matrix_row_desc:
        for j in range(1, len(options) + 1):
            ws_sum.cell(
                row=row_desc,
                column=1 + j,
                value=f"='Matrix'!{get_column_letter(2 + j)}{matrix_row_desc}",
            ).alignment = LEFT_SUM

    # Score row: SAW Score from Weights & SAW (total_row), displayed as %
    for j in range(1, len(options) + 1):
        colW = 4 + j  # E.. for SAW per option (same columns as NormScore headers)
        colS = 1 + j
        formula = f"='Weights & SAW'!{get_column_letter(colW)}{total_row}"
        c = ws_sum.cell(row=row_score, column=colS, value=formula)
        c.number_format = "0%"  # show as 0–100%
        c.alignment = CENTER_SUM

    # Conditional formatting on Score row (same bands/colors as elsewhere)
    bands_pct = [
        (0, 0.20, "FFC000"),
        (0.20, 0.40, "FFFF66"),
        (0.40, 0.60, "CCFF66"),
        (0.60, 0.80, "78FE66"),
        (0.80, 1.01, "19CB01"),
    ]
    for j in range(1, len(options) + 1):
        addr = f"{get_column_letter(1 + j)}{row_score}"
        for lo, hi, hexcol in bands_pct:
            expr = f"=AND({addr}>={lo},{addr}<{hi})"
            dxf = DifferentialStyle(fill=PatternFill(fill_type="solid", start_color=hexcol, end_color=hexcol))
            ws_sum.conditional_formatting.add(addr, Rule(type="expression", dxf=dxf, formula=[expr]))

    # Summary row: Pros/Cons pulled from Matrix Advantages/Disadvantages
    adv_row = find_matrix_row("advantages")
    dis_row = find_matrix_row("disadvantages")

    for j in range(1, len(options) + 1):
        adv_ref = f"'Matrix'!{get_column_letter(2 + j)}{adv_row}" if adv_row else '""'
        dis_ref = f"'Matrix'!{get_column_letter(2 + j)}{dis_row}" if dis_row else '""'

        if adv_row and dis_row:
            formula = (
                f'=IF(OR(LEN({adv_ref})>0,LEN({dis_ref})>0),'
                f'IF(LEN({adv_ref})>0,"Pros:"&CHAR(10)&{adv_ref},"")'
                f'&IF(AND(LEN({adv_ref})>0,LEN({dis_ref})>0),CHAR(10)&CHAR(10),"")'
                f'&IF(LEN({dis_ref})>0,"Cons:"&CHAR(10)&{dis_ref},""),"")'
            )
        elif adv_row:
            formula = f'=IF(LEN({adv_ref})>0,"Pros:"&CHAR(10)&{adv_ref},"")'
        elif dis_row:
            formula = f'=IF(LEN({dis_ref})>0,"Cons:"&CHAR(10)&{dis_ref},"")'
        else:
            formula = '=""'

        ws_sum.cell(row=row_summary, column=1 + j, value=formula).alignment = LEFT_SUM

    # Borders & widths
    last_col_sum = 1 + len(options)
    for rr in range(row_ill, row_summary + 1):
        for cc in range(1, last_col_sum + 1):
            left_b = THICK if cc == 1 else THIN
            right_b = THICK if cc == last_col_sum else THIN
            top_b = THICK if rr == row_ill else THIN
            bottom_b = THICK if rr == row_summary else THIN
            if cc >= 2 and cc < last_col_sum:
                left_b = MEDIUM
            ws_sum.cell(row=rr, column=cc).border = Border(left=left_b, right=right_b, top=top_b, bottom=bottom_b)

    ws_sum.column_dimensions["A"].width = 18
    for col in range(2, last_col_sum + 1):
        ws_sum.column_dimensions[get_column_letter(col)].width = 35
    ws_sum.freeze_panes = "B4"

    # Summary CBA fixed row heights
    ws_sum.row_dimensions[1].height = 28   # Title
    ws_sum.row_dimensions[2].height = 22   # Project info
    ws_sum.row_dimensions[3].height = 10   # Spacer
    ws_sum.row_dimensions[row_ill].height = 60
    ws_sum.row_dimensions[row_opt].height = 22
    ws_sum.row_dimensions[row_desc].height = 90
    ws_sum.row_dimensions[row_score].height = 22
    ws_sum.row_dimensions[row_summary].height = 160

    # ---- Save to bytes----
    out_name = safe_name(f"TEG CBA Matrix-{purpose}-{project_name}-{date.today():%m%d%Y}") + ".xlsx"
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue(), out_name
